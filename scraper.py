import sys
import json
import os
import asyncio
import re
from datetime import datetime
from urllib.parse import urlparse
import pandas as pd
from tenacity import retry, stop_after_attempt, wait_exponential

if sys.platform == 'win32':
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

from playwright_stealth import Stealth
from playwright.async_api import async_playwright
import extruct
from w3lib.html import get_base_url

# Pre-deployment check
if os.environ.get('STREAMLIT_RUNTIME_CHECK') or not os.name == 'nt':
    if not os.path.exists("/home/adminuser/.cache/ms-playwright"):
        os.system("playwright install chromium > /dev/null 2>&1")


# ---------------------------------------------------------------------------
# FILENAME HELPER
# ---------------------------------------------------------------------------

def url_to_filename(url: str) -> str:
    parsed = urlparse(url)
    hostname = parsed.hostname or ""
    hostname = re.sub(r"^www\.", "", hostname)
    hostname_stem = hostname.rsplit(".", 1)[0]
    path = parsed.path.strip("/")
    path = re.sub(r"/+", "_", path)
    hostname_stem = re.sub(r"[^\w\-]", "", hostname_stem)
    path = re.sub(r"[^\w\-]", "", path)
    date_str = datetime.now().strftime("%m%d%Y")
    parts = [p for p in [hostname_stem, path] if p]
    return "_".join(parts) + f"_{date_str}"


# ---------------------------------------------------------------------------
# JS VARIABLE EXTRACTION
# ---------------------------------------------------------------------------

def extract_js_variables(html: str, source_url: str) -> list:
    rows = []
    extractors = [
        ("js:utag_data",   r'utag_data\s*=\s*(\{.*?\});'),
        ("js:priceObject", r'(?:let|var|const)\s+priceObject\s*=\s*(\{.*?\});'),
        ("js:dataLayer",   r'dataLayer\.push\((\{.*?\})\)'),
    ]
    for syntax_label, pattern in extractors:
        matches = re.findall(pattern, html, re.DOTALL)
        for raw in matches:
            try:
                obj = json.loads(raw.strip())
            except json.JSONDecodeError:
                continue
            if not isinstance(obj, dict):
                continue
            schema_type = obj.get("@type") or obj.get("page_type") or obj.get("event") or "unknown"
            for key, value in obj.items():
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, ensure_ascii=False)
                rows.append({
                    "source_url": source_url,
                    "syntax": syntax_label,
                    "schema_type": str(schema_type),
                    "key": key,
                    "value": str(value),
                })
    return rows


# ---------------------------------------------------------------------------
# FLATTEN STRUCTURED METADATA
# ---------------------------------------------------------------------------

def flatten_data(data: dict, source_url: str) -> list:
    rows = []
    for syntax, items in data.items():
        if not isinstance(items, list):
            continue
        for item in items:
            if not isinstance(item, dict):
                continue
            schema_type = item.get("@type") or item.get("type") or "unknown"
            if isinstance(schema_type, list):
                schema_type = ", ".join(schema_type)
            for key, value in item.items():
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, ensure_ascii=False)
                rows.append({
                    "source_url": source_url,
                    "syntax": syntax,
                    "schema_type": schema_type,
                    "key": key,
                    "value": str(value),
                })
    return rows


# ---------------------------------------------------------------------------
# GEO / UCP ANALYZER
# ---------------------------------------------------------------------------

def _json_ld_rows(rows):
    return [r for r in rows if r.get("syntax") == "json-ld"]

def _values_for_key(rows, key):
    return [r["value"] for r in rows if r["key"] == key]

def _has_type(rows, type_name):
    """
    Check for a schema @type in three places:
      1. The schema_type column (standard flat rows)
      2. Inside any @graph JSON blob (Yoast/WordPress @graph pattern)
      3. Inside any large JSON blob value (catches SAP Hybris / Carhartt pattern
         where extruct stores the whole JSON-LD array as a single value string
         because the page uses http://schema.org context or <script id="json-ld">)
    """
    # Standard check
    if any(type_name.lower() in r.get("schema_type", "").lower() for r in rows):
        return True
    # Graph-aware check
    for r in rows:
        if r.get("key") == "@graph":
            try:
                graph = json.loads(r["value"])
                if isinstance(graph, list):
                    for item in graph:
                        if isinstance(item, dict):
                            t = item.get("@type", "")
                            if isinstance(t, list):
                                if any(type_name.lower() in x.lower() for x in t):
                                    return True
                            elif type_name.lower() in str(t).lower():
                                return True
            except (json.JSONDecodeError, TypeError):
                continue
    # Raw blob scan — catches cases where extruct stored a whole JSON-LD
    # array/object as a single value string (http: context, id="json-ld", etc.)
    for r in rows:
        val = r.get("value", "")
        if len(val) > 200 and f'"@type"' in val and type_name in val:
            try:
                parsed = json.loads(val)
                objs = parsed if isinstance(parsed, list) else [parsed]
                for obj in objs:
                    if isinstance(obj, dict):
                        t = obj.get("@type", "")
                        types = t if isinstance(t, list) else [t]
                        if any(type_name.lower() in x.lower() for x in types):
                            return True
                        # Also check one level deep (nested sub-objects)
                        for v in obj.values():
                            if isinstance(v, dict):
                                sub_t = v.get("@type", "")
                                sub_types = sub_t if isinstance(sub_t, list) else [sub_t]
                                if any(type_name.lower() in x.lower()
                                       for x in sub_types):
                                    return True
            except (json.JSONDecodeError, TypeError):
                continue
    return False

def _graph_values_for_type_key(rows, type_name, key_name):
    """
    Extract values from inside @graph items matching type_name and key_name.
    Used when the data is nested inside a @graph blob rather than flat rows.
    """
    values = []
    for r in rows:
        if r.get("key") == "@graph":
            try:
                graph = json.loads(r["value"])
                if isinstance(graph, list):
                    for item in graph:
                        if isinstance(item, dict):
                            t = item.get("@type", "")
                            types = t if isinstance(t, list) else [t]
                            if any(type_name.lower() in x.lower() for x in types):
                                if key_name in item:
                                    val = item[key_name]
                                    values.append(
                                        json.dumps(val) if isinstance(val, (dict, list))
                                        else str(val)
                                    )
            except (json.JSONDecodeError, TypeError):
                continue
    return values

def _all_schema_values_str(rows):
    """Concatenate all values for regex scanning (IP detection, etc.)."""
    return " ".join(str(r.get("value", "")) for r in rows)


def analyze(rows, source_url=""):
    """
    Run all GEO/UCP readiness checks against scraped rows.
    Returns a list of finding dicts sorted by severity then priority.
    """
    findings = []
    jld = _json_ld_rows(rows)

    def add(priority, category, status, issue, fix, effort, impact):
        findings.append({
            "Priority": priority,
            "Category": category,
            "Status":   status,
            "Issue":    issue,
            "Fix":      fix,
            "Effort":   effort,
            "Impact":   impact,
        })

    # ── 1. Schema @context protocol ─────────────────────────────────────────
    ctx_values = _values_for_key(rows, "@context")
    if any("http://" in v for v in ctx_values):
        add(4, "Schema Context", "WARN",
            "One or more @context values use http://schema.org instead of https://schema.org.",
            "Update all @context values to 'https://schema.org'. The https form is canonical "
            "per schema.org documentation and avoids strict-parser warnings in audit tools.",
            "Low", "Low")
    elif ctx_values:
        add(5, "Schema Context", "PASS",
            "@context uses https://schema.org correctly.", "", "—", "—")

    # ── 2. Core schema type present ──────────────────────────────────────────
    has_product = _has_type(jld, "Product")
    has_service = _has_type(jld, "Service")
    has_local   = _has_type(jld, "LocalBusiness")

    if not has_product and not has_service and not has_local:
        add(1, "Core Schema", "FAIL",
            "No Product, Service, or LocalBusiness schema detected. Without this, AI engines "
            "have no structured signal about what this page represents and cannot cite or "
            "recommend it for relevant queries.",
            "Add a JSON-LD <script> block in the page <head> with @type: Product (product "
            "pages), Service (service pages), or LocalBusiness (company/location pages). "
            "This is the non-negotiable foundation for all other schema to be useful.",
            "Medium", "High")
    else:
        found = [t for t, b in [("Product", has_product), ("Service", has_service),
                                  ("LocalBusiness", has_local)] if b]
        add(5, "Core Schema", "PASS",
            f"Core schema type(s) found: {', '.join(found)}.", "", "—", "—")

    # ── 3. GTIN / UPC (product pages only) ──────────────────────────────────
    if has_product:
        gtin_keys = {"gtin", "gtin8", "gtin12", "gtin13", "gtin14"}
        found_gtins = [r for r in jld if r.get("key", "").lower() in gtin_keys
                       and r.get("value", "").strip()]
        # Also check graph
        for k in gtin_keys:
            found_gtins += _graph_values_for_type_key(jld, "Product", k)

        if not found_gtins:
            add(1, "Product Identification", "FAIL",
                "No GTIN (UPC/EAN) found in Product schema. GTIN is the primary identifier "
                "AI systems and Google Shopping use to match this product across multiple "
                "sellers and product databases. Without it the product is opaque to "
                "cross-platform matching and will be excluded from many AI commerce surfaces.",
                "Add the product UPC or EAN as 'gtin12' or 'gtin13' in the Product JSON-LD "
                "block. Example: \"gtin13\": \"00012345678905\"  "
                "Find the UPC on the product packaging or manufacturer spec sheet.",
                "Low", "High")
        else:
            val = found_gtins[0] if isinstance(found_gtins[0], str) else found_gtins[0].get("value","")
            add(5, "Product Identification", "PASS",
                f"GTIN found: {str(val)[:40]}", "", "—", "—")

    # ── 4. Price validity (product pages) ───────────────────────────────────
    if has_product:
        price_vals = _values_for_key(rows, "price")
        # Also look inside offers/priceSpecification in graph
        offer_graph = _graph_values_for_type_key(jld, "Offer", "price")
        price_vals += offer_graph

        # Also check for lowPrice / highPrice inside AggregateOffer blobs —
        # Carhartt, Magento, and SAP Hybris frequently emit:
        #   {"@type": "AggregateOffer", "lowPrice": 149.99, "highPrice": 164.99, ...}
        # stored as a single JSON value rather than flat rows.
        # Also handles price inside priceSpecification (WooCommerce / Happy Howie's pattern):
        #   {"@type": "Offer", "priceSpecification": [{"price": "6.79", ...}]}
        if not price_vals:
            agg_price_keys = {"lowprice", "highprice", "price"}
            for r in jld:
                val_str = r.get("value", "")
                if len(val_str) > 10 and '"@type"' in val_str:
                    try:
                        val_obj = json.loads(val_str)
                        objs = val_obj if isinstance(val_obj, list) else [val_obj]
                        for obj in objs:
                            if not isinstance(obj, dict):
                                continue
                            # Check direct keys on the offer object
                            for k, v in obj.items():
                                if k.lower() in agg_price_keys and v not in (None, "", 0, "0"):
                                    price_vals.append(str(v))
                            # Check one level deep (offers nested inside Product)
                            for sub_val in obj.values():
                                if isinstance(sub_val, dict):
                                    for k, v in sub_val.items():
                                        if k.lower() in agg_price_keys and v not in (None, "", 0, "0"):
                                            price_vals.append(str(v))
                                # Also check lists of sub-objects (priceSpecification is a list)
                                elif isinstance(sub_val, list):
                                    for sub_item in sub_val:
                                        if isinstance(sub_item, dict):
                                            for k, v in sub_item.items():
                                                if k.lower() in agg_price_keys and v not in (None, "", 0, "0"):
                                                    price_vals.append(str(v))
                    except (json.JSONDecodeError, TypeError):
                        continue
                if price_vals:
                    break

        if not price_vals:
            add(2, "Pricing", "WARN",
                "No price found in schema. If this page shows a public price it should be "
                "in the Offer schema. If pricing requires login, omit the price fields "
                "entirely — do not default to 0.",
                "Add 'price' and 'priceCurrency' to the Offer block if a public price exists. "
                "For login-gated pricing, remove price and priceCurrency from the Offer "
                "block and keep only availability and seller.",
                "Low", "Medium")
        else:
            zero_price = any(v.strip() in ("0", "0.0", "0.00") for v in price_vals)
            if zero_price:
                add(1, "Pricing", "FAIL",
                    "Price is set to 0 in Offer schema. Google and AI commerce systems treat "
                    "price: 0 as either a free product or broken data. This causes the product "
                    "to be excluded from Shopping surfaces and AI price comparison responses.",
                    "Replace price: 0 with the real price, or remove price and priceCurrency "
                    "entirely if the page does not show a public price. "
                    "A missing price field is always better than a zero price.",
                    "Low", "High")
            else:
                add(5, "Pricing", "PASS",
                    f"Valid price found: {price_vals[0]}", "", "—", "—")

        # priceValidUntil
        pvu = _values_for_key(rows, "priceValidUntil")
        pvu += _graph_values_for_type_key(jld, "Offer", "priceValidUntil")
        if not pvu:
            add(3, "Pricing", "WARN",
                "No priceValidUntil date in Offer schema. AI agents treat pricing without "
                "an expiry date as potentially stale and may deprioritize it.",
                "Add 'priceValidUntil' with a future date (e.g. '2027-12-31') to the "
                "Offer block.",
                "Low", "Medium")
        else:
            add(5, "Pricing", "PASS",
                f"priceValidUntil found: {pvu[0][:10]}", "", "—", "—")

    # ── 5. Availability URI correctness ──────────────────────────────────────
    avail_vals = _values_for_key(rows, "availability")
    avail_vals += _graph_values_for_type_key(jld, "Offer", "availability")
    if avail_vals:
        if any("http://" in v for v in avail_vals):
            add(2, "Schema Quality", "WARN",
                "availability uses http://schema.org/InStock instead of "
                "https://schema.org/InStock. Google flags this specific field as a "
                "structured data warning in Search Console.",
                "Change 'http://schema.org/InStock' to 'https://schema.org/InStock' "
                "in the Offer block. This is a one-character protocol change.",
                "Low", "Medium")
        else:
            add(5, "Schema Quality", "PASS",
                f"availability uses correct https URI: {avail_vals[0][:50]}", "", "—", "—")

    # ── 6. AggregateRating ───────────────────────────────────────────────────
    # Three-layer check:
    #   (a) Standard: AggregateRating as a top-level schema_type row
    #   (b) Graph-aware: inside a @graph blob (handled by _has_type already)
    #   (c) Nested: aggregateRating key inside a Product/LocalBusiness flat row
    #       — this is how Carhartt, Magento, and many custom platforms emit it:
    #         the entire Product is one JSON-LD object, and aggregateRating is
    #         a nested field rather than a separate top-level block.
    has_agg = _has_type(jld, "AggregateRating")

    if not has_agg:
        for r in jld:
            if r.get("key", "").lower() == "aggregaterating":
                val_str = r.get("value", "")
                # Could be a JSON object string or a plain value
                try:
                    val_obj = json.loads(val_str)
                    if isinstance(val_obj, dict) and (
                        val_obj.get("@type", "").lower() == "aggregaterating"
                        or "ratingValue" in val_obj
                        or "ratingvalue" in val_obj
                    ):
                        has_agg = True
                        break
                except (json.JSONDecodeError, TypeError):
                    # Key presence alone is a strong signal even if not parseable
                    if val_str.strip():
                        has_agg = True
                        break

    if has_agg:
        rv = (_values_for_key(rows, "ratingValue") or
              _graph_values_for_type_key(jld, "AggregateRating", "ratingValue"))
        rc = (_values_for_key(rows, "reviewCount") or
              _graph_values_for_type_key(jld, "AggregateRating", "reviewCount"))
        # Also extract from nested aggregateRating key if flat lookup missed it
        if not rv or not rc:
            for r in jld:
                if r.get("key", "").lower() == "aggregaterating":
                    try:
                        val_obj = json.loads(r.get("value", ""))
                        if isinstance(val_obj, dict):
                            if not rv and val_obj.get("ratingValue"):
                                rv = [str(val_obj["ratingValue"])]
                            if not rc:
                                rc = [str(val_obj.get("reviewCount") or
                                          val_obj.get("ratingCount", "?"))]
                    except (json.JSONDecodeError, TypeError):
                        pass
        add(5, "Social Proof", "PASS",
            f"AggregateRating found: {rv[0] if rv else '?'} stars, "
            f"{rc[0] if rc else '?'} reviews.", "", "—", "—")
    else:
        add(2, "Social Proof", "FAIL",
            "No AggregateRating schema found. Star ratings are a primary signal AI systems "
            "use when recommending products or services. A competitor page with even a small "
            "number of structured reviews will outrank this page for evaluative queries "
            "like 'best', 'recommended', or 'highest rated'.",
            "Add an AggregateRating block nested inside the Product or LocalBusiness schema "
            "with ratingValue, reviewCount, bestRating (5), and worstRating (1). "
            "Only include real ratings — never fabricate review data.",
            "Medium", "High")

    # ── 7. FAQPage schema ────────────────────────────────────────────────────
    if _has_type(jld, "FAQPage") or _has_type(jld, "Question"):
        add(5, "GEO Content", "PASS",
            "FAQPage schema found.", "", "—", "—")
    else:
        add(2, "GEO Content", "FAIL",
            "No FAQPage schema found. FAQ schema is one of the highest-impact GEO additions "
            "because it gives AI systems machine-readable, directly citable answers to the "
            "exact questions users ask. Pages with FAQ schema are significantly more likely "
            "to be quoted verbatim in AI-generated responses.",
            "Add a FAQPage JSON-LD block with 4-8 real questions and detailed answers "
            "relevant to this product or service. Base questions on what customers actually "
            "ask — product use, compatibility, certifications, installation, warranty, etc. "
            "Avoid marketing language; write as if answering a direct customer question.",
            "Medium", "High")

    # ── 8. BreadcrumbList ────────────────────────────────────────────────────
    # Two-layer check:
    #   (a) Standard via _has_type (covers https: context and @graph patterns)
    #   (b) Raw value scan — catches cases where extruct skips the block because
    #       the page uses http://schema.org context (Carhartt, SAP Hybris pattern)
    #       or where <script id="json-ld"> is used instead of a plain type attr.
    has_breadcrumb = _has_type(jld, "BreadcrumbList")

    if not has_breadcrumb:
        # Check if any flat row value contains BreadcrumbList anywhere
        for r in rows:
            val = r.get("value", "")
            if "BreadcrumbList" in val:
                has_breadcrumb = True
                break

    if has_breadcrumb:
        add(5, "Navigation", "PASS",
            "BreadcrumbList schema found.", "", "—", "—")
    else:
        add(3, "Navigation", "FAIL",
            "No BreadcrumbList schema. Breadcrumbs help AI agents understand where this "
            "page sits in the site hierarchy, improving category-level query matching "
            "and enabling more accurate product/service classification.",
            "Add a BreadcrumbList JSON-LD block with one ListItem per navigation level. "
            "Each ListItem needs: position (integer), name (label), and item (full URL). "
            "The final item (current page) can omit the item URL.",
            "Low", "Medium")

    # ── 9. Organization / entity identity ───────────────────────────────────
    # Three-layer check:
    #   (a) Standalone Organization type, OR any recognized schema.org subtype
    #       of Organization/LocalBusiness (ProfessionalService, HomeAndConstructionBusiness,
    #       Restaurant, Store, Corporation, NGO, etc. all inherit from Organization —
    #       a literal-string match on "Organization"/"LocalBusiness" alone misses these,
    #       which is what caused the false FAIL on T2 Digital's ProfessionalService block).
    #   (b) @graph pattern (handled by _has_type already)
    #   (c) Nested brand node inside Product — many manufacturer sites (Carhartt,
    #       Nike, etc.) embed brand as {"@type":"Brand","name":"X"} inside
    #       the Product block rather than a separate Organization object.
    #       A named Brand node is sufficient to pass entity identity for product pages.
    ORG_SUBTYPES = {
        "organization", "localbusiness", "professionalservice", "corporation",
        "ngo", "store", "restaurant", "homeandconstructionbusiness",
        "medicalbusiness", "financialservice", "foodestablishment",
        "lodgingbusiness", "automotivebusiness", "legalservice",
        "sportsorganization", "educationalorganization", "governmentorganization",
        "performinggroup", "newsmediaorganization", "onlinebusiness",
        "consortium", "ngoorganization", "ngo",
    }

    def _has_org_type(rows_to_check):
        for r in rows_to_check:
            stype = r.get("schema_type", "").lower()
            if any(sub in stype for sub in ORG_SUBTYPES):
                return True
        # Raw blob scan — same pattern as _has_type's third layer, for cases
        # where the whole JSON-LD object is stored as one value string
        for r in rows_to_check:
            val = r.get("value", "")
            if len(val) > 50 and '"@type"' in val:
                try:
                    parsed = json.loads(val)
                    objs = parsed if isinstance(parsed, list) else [parsed]
                    for obj in objs:
                        if isinstance(obj, dict):
                            t = obj.get("@type", "")
                            types = [t] if isinstance(t, str) else (t if isinstance(t, list) else [])
                            if any(str(x).lower() in ORG_SUBTYPES for x in types):
                                return True
                except (json.JSONDecodeError, TypeError):
                    continue
        return False

    has_org = _has_org_type(jld)

    if not has_org and has_product:
        # Check for a brand node with @type Brand or Organization nested in Product
        for r in jld:
            if r.get("key", "").lower() == "brand":
                val_str = r.get("value", "")
                try:
                    val_obj = json.loads(val_str)
                    if isinstance(val_obj, dict) and val_obj.get("name"):
                        has_org = True
                        break
                except (json.JSONDecodeError, TypeError):
                    # Plain string brand name also counts
                    if val_str.strip():
                        has_org = True
                        break

    if not has_org:
        add(2, "Entity Identity", "FAIL",
            "No Organization or LocalBusiness schema found. AI agents need a structured "
            "entity node to confidently identify who makes or sells this product/service. "
            "Without it the brand is poorly disambiguated across the web.",
            "Add an Organization block with at minimum: name, url, logo (ImageObject), "
            "and sameAs (array of social/directory profile URLs). For local businesses "
            "use LocalBusiness and add address, telephone, and areaServed.",
            "Low", "High")
    else:
        # Check sameAs — in flat rows or inside @graph Organization item
        same_as = (_values_for_key(rows, "sameAs") or
                   _graph_values_for_type_key(jld, "Organization", "sameAs"))
        if same_as:
            add(5, "Entity Identity", "PASS",
                "Organization with sameAs social links found.", "", "—", "—")
        else:
            add(3, "Entity Identity", "WARN",
                "Organization schema present but no sameAs links found. sameAs links "
                "to social and directory profiles are how AI engines confirm entity "
                "identity across the web and avoid confusing similarly-named businesses.",
                "Add a sameAs array to the Organization block with URLs for each "
                "active social or directory profile: Facebook, LinkedIn, Instagram, "
                "Yelp, Google Business Profile, etc.",
                "Low", "Medium")

    # ── 10. Raw IP address in schema URLs ───────────────────────────────────
    all_vals = _all_schema_values_str(rows)
    if re.search(r'https?://\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}/', all_vals):
        add(1, "Technical — URLs", "FAIL",
            "A raw IP address was found in schema URLs (almost always the Organization "
            "logo). This is a leftover from a server migration where the IP was in the "
            "database when the logo was last saved. Google and AI crawlers may refuse to "
            "fetch assets from raw IPs with no SSL domain trust, leaving the brand "
            "without a recognized logo in AI responses.",
            "Update the logo URL in your SEO plugin settings to use the proper https:// "
            "domain path. In WordPress/Yoast: SEO > Settings > Site Representation > "
            "click the logo thumbnail and re-select it from the Media Library. Save. "
            "Verify fix by searching page source for the old IP.",
            "Low", "High")

    # ── 11. og:type mismatch on product pages ────────────────────────────────
    # og:type comes from Open Graph meta tags, which extruct surfaces as RDFa
    # rows (key: http://ogp.me/ns#type), NOT as JSON-LD. We check RDFa only —
    # the JSON-LD @type: Product is a separate, correct field and is not in scope
    # here. Only flag when we can confirm the RDFa og:type value is "article".
    og_type_vals = []
    for r in rows:
        key = r.get("key", "")
        syntax = r.get("syntax", "")
        # RDFa pattern from extruct: key = http://ogp.me/ns#type
        if "ogp.me" in key and key.endswith("#type") and syntax == "rdfa":
            raw = r.get("value", "")
            # extruct wraps RDFa values as [{"@value": "article"}]
            try:
                parsed = json.loads(raw)
                if isinstance(parsed, list):
                    for item in parsed:
                        if isinstance(item, dict) and "@value" in item:
                            og_type_vals.append(str(item["@value"]))
                else:
                    og_type_vals.append(str(raw))
            except (json.JSONDecodeError, TypeError):
                og_type_vals.append(str(raw))
        # Plain og:type key (some parsers flatten it this way)
        elif key == "og:type":
            og_type_vals.append(str(r.get("value", "")))

    if has_product and any("article" in v.lower() for v in og_type_vals):
        add(2, "Open Graph", "WARN",
            "The og:type Open Graph meta tag is set to 'article' on a product page. "
            "This is a separate field from your JSON-LD schema — it lives in the page's "
            "<head> as an Open Graph meta tag. Social platforms (Facebook, LinkedIn) and "
            "some AI crawlers use og:type to classify content; 'article' signals editorial "
            "content rather than a purchasable item, which can affect commerce surface "
            "eligibility and social sharing previews.",
            "Change the og:type meta tag value to 'product'. This is NOT in your JSON-LD "
            "block — look in your CMS or SEO plugin settings instead. "
            "In WooCommerce with Yoast: SEO > Social > Facebook > set type to Product, "
            "or check the product's individual Yoast Social panel. "
            "In RankMath: Schema > Open Graph Type > Product.",
            "Low", "Medium")

    # ── 12. shippingDetails (product pages) ─────────────────────────────────
    if has_product:
        has_shipping = (any(r.get("key") == "shippingDetails" for r in rows) or
                        bool(_graph_values_for_type_key(jld, "Offer", "shippingDetails")))
        if not has_shipping:
            add(3, "Commerce Completeness", "WARN",
                "No shippingDetails in Offer schema. Shipping cost and delivery time are "
                "primary decision factors for AI agents doing purchase comparison. Without "
                "them the product is invisible to 'best option including shipping' queries.",
                "Add an OfferShippingDetails node inside the Offer block with: "
                "shippingRate (MonetaryAmount with value and currency) and "
                "deliveryTime (ShippingDeliveryTime with businessDays min/max).",
                "Medium", "Medium")

    # ── 13. Return policy (product pages) ───────────────────────────────────
    if has_product:
        has_return = any("returnpolicy" in r.get("key", "").lower() or
                         "merchantreturn" in r.get("key", "").lower()
                         for r in rows)
        if not has_return:
            add(3, "Commerce Completeness", "WARN",
                "No hasMerchantReturnPolicy found. AI agents comparing purchase options "
                "factor in return policies as a trust and convenience signal.",
                "Add a MerchantReturnPolicy node to the Offer block with: "
                "returnPolicyCategory (e.g. https://schema.org/MerchantReturnFiniteReturnWindow) "
                "and merchantReturnDays (integer number of days).",
                "Medium", "Low")

    # ── 14. Seller node in Offer (product pages) ─────────────────────────────
    if has_product:
        has_seller = (any(r.get("key") == "seller" for r in rows) or
                      bool(_graph_values_for_type_key(jld, "Offer", "seller")))
        if not has_seller:
            add(3, "Commerce Completeness", "WARN",
                "No seller node in Offer schema. For distributor or retailer pages where "
                "the seller differs from the manufacturer, AI agents need the seller "
                "identified explicitly to answer 'who sells this?' correctly.",
                "Add a seller node inside the Offer block: "
                "{\"@type\": \"Organization\", \"name\": \"Your Company Name\", "
                "\"url\": \"https://yoursite.com\"}",
                "Low", "Medium")
        else:
            add(5, "Commerce Completeness", "PASS",
                "seller node found in Offer.", "", "—", "—")

    # ── 15. brand node on Product ─────────────────────────────────────────────
    if has_product:
        has_brand = (any(r.get("key") == "brand" for r in rows) or
                     bool(_graph_values_for_type_key(jld, "Product", "brand")))
        if not has_brand:
            add(3, "Product Identification", "WARN",
                "No brand node on Product schema. For branded consumer products, the brand "
                "connection matters — it is how AI systems confirm 'this product is made by X' "
                "rather than just sold by X.",
                "Add a brand node to the Product block: "
                "{\"@type\": \"Brand\", \"name\": \"Brand Name\", \"url\": \"https://brand.com\"}",
                "Low", "Medium")
        else:
            add(5, "Product Identification", "PASS",
                "brand node found on Product.", "", "—", "—")

    # ── 16. additionalProperty for key product attributes ────────────────────
    if has_product:
        has_props = any(r.get("key") == "additionalProperty" for r in rows)
        if not has_props:
            add(4, "Product Detail", "WARN",
                "No additionalProperty nodes found. Product-specific attributes like "
                "weight, dimensions, material, target species (for pet products), "
                "voltage/amperage (for electrical), or thread size (for fittings) are "
                "not machine-readable. AI agents answering spec questions must parse prose.",
                "Add additionalProperty nodes using PropertyValue for each key spec: "
                "{\"@type\": \"PropertyValue\", \"name\": \"Net Weight\", \"value\": \"7 oz\"}. "
                "Include the attributes customers most frequently ask about.",
                "Medium", "Medium")

    # ── 17. areaServed (local / service pages) ───────────────────────────────
    if has_local or has_service:
        has_area = (any(r.get("key") == "areaServed" for r in rows) or
                    bool(_graph_values_for_type_key(jld, "LocalBusiness", "areaServed")) or
                    bool(_graph_values_for_type_key(jld, "Service", "areaServed")))
        if not has_area:
            add(1, "Local Discovery", "FAIL",
                "No areaServed on LocalBusiness or Service schema. This is the primary field "
                "AI engines use to match 'near me' and location-specific queries. Without it "
                "the business will not appear in AI responses to local service queries "
                "regardless of all other schema quality.",
                "Add areaServed to LocalBusiness or Service listing each city or county "
                "served as its own object: {\"@type\": \"City\", \"name\": \"Lafayette\"}. "
                "List every meaningful service area — AI matching is exact.",
                "Low", "High")
        else:
            add(5, "Local Discovery", "PASS",
                "areaServed found.", "", "—", "—")

    # ── 18. openingHoursSpecification (local business) ───────────────────────
    if has_local:
        has_hours = any("openinghours" in r.get("key", "").lower() for r in rows)
        if not has_hours:
            add(3, "Local Business", "WARN",
                "No openingHoursSpecification found. AI agents checking whether a "
                "business is currently open or available rely on this field for "
                "time-sensitive queries.",
                "Add openingHoursSpecification to the LocalBusiness block with "
                "dayOfWeek, opens, and closes for each schedule. Add a separate entry "
                "for any 24/7 emergency lines with opens: '00:00' closes: '23:59'.",
                "Low", "Medium")

    # ── 19. Certification / credentials (service businesses) ─────────────────
    if has_local or has_service:
        has_cert = (_has_type(jld, "Certification") or
                    any("credential" in r.get("key", "").lower() or
                        "certification" in r.get("key", "").lower()
                        for r in rows))
        if not has_cert:
            add(3, "Credentials", "WARN",
                "No Certification or credential schema found. For licensed service "
                "businesses (arborists, electricians, contractors, etc.), structured "
                "credential data is a meaningful differentiator — AI engines can filter "
                "results by 'certified' or 'licensed' when users request it, and will "
                "prefer businesses with verifiable credential schema.",
                "Add a hasCredential block to the Organization/LocalBusiness node: "
                "{\"@type\": \"EducationalOccupationalCredential\", "
                "\"name\": \"ISA Certified Arborist #IN-3476A\", "
                "\"recognizedBy\": {\"@type\": \"Organization\", "
                "\"name\": \"International Society of Arboriculture\"}}",
                "Low", "Medium")

    # ── 20. dateModified freshness ───────────────────────────────────────────
    modified_vals = (_values_for_key(rows, "dateModified") or
                     _graph_values_for_type_key(jld, "WebPage", "dateModified"))
    if not modified_vals:
        add(4, "Content Freshness", "WARN",
            "No dateModified found in schema. Pages with a recent dateModified signal "
            "receive priority for time-sensitive queries in some AI ranking systems.",
            "Add dateModified in ISO 8601 format to the WebPage schema block, "
            "reflecting the date of the last meaningful content update. "
            "Example: \"dateModified\": \"2026-06-12T10:00:00+00:00\"",
            "Low", "Low")
    else:
        try:
            raw_date = modified_vals[0][:25]
            mod_date = datetime.fromisoformat(raw_date.replace("Z", "+00:00"))
            age_days = (datetime.now(mod_date.tzinfo) - mod_date).days
            if age_days > 180:
                add(3, "Content Freshness", "WARN",
                    f"dateModified is {age_days} days old ({modified_vals[0][:10]}). "
                    "Content not updated in 6+ months may be deprioritized for "
                    "freshness-sensitive queries.",
                    "Update page content and refresh the dateModified value to the "
                    "date of the last meaningful change.",
                    "Low", "Medium")
            else:
                add(5, "Content Freshness", "PASS",
                    f"dateModified is recent: {modified_vals[0][:10]} ({age_days} days ago).",
                    "", "—", "—")
        except (ValueError, TypeError):
            pass

    # ── 21. Description quality ──────────────────────────────────────────────
    desc_vals = (_values_for_key(rows, "description") or
                 _graph_values_for_type_key(jld, "Product", "description") or
                 _graph_values_for_type_key(jld, "WebPage", "description"))
    if not desc_vals:
        add(2, "Content Quality", "WARN",
            "No description found in schema. AI systems use the structured description "
            "as a primary source when generating citations and summaries.",
            "Add a description field to the Product, Service, or WebPage schema block. "
            "Write in full prose sentences — not bullet points or keyword lists. "
            "Aim for 150-300 characters covering name, key benefit, and use case.",
            "Low", "Medium")
    else:
        desc = desc_vals[0]
        if re.search(r'lorem ipsum', desc, re.IGNORECASE):
            add(1, "Content Quality", "FAIL",
                "Lorem ipsum placeholder text detected in schema description. "
                "Crawlers and AI systems treat this as low-quality or unfinished content, "
                "which harms trust signals across the entire domain.",
                "Replace all lorem ipsum placeholder text with real content describing "
                "this product or service before publishing.",
                "Low", "High")
        elif len(desc) < 80:
            add(3, "Content Quality", "WARN",
                f"Description is very short ({len(desc)} characters). AI engines prefer "
                "enough detail to generate meaningful citations.",
                "Expand the description to at least 150 characters of plain prose. "
                "Include the product/service name, primary benefit, and key use case.",
                "Low", "Medium")
        else:
            add(5, "Content Quality", "PASS",
                f"Description found ({len(desc)} chars).", "", "—", "—")

    # ── 22. SearchAction / UCP discoverability ───────────────────────────────
    has_search = (any("SearchAction" in r.get("value", "") or
                      r.get("key") == "potentialAction" for r in rows) or
                  bool(_graph_values_for_type_key(jld, "WebSite", "potentialAction")))
    if has_search:
        add(5, "UCP Discoverability", "PASS",
            "SearchAction potentialAction found.", "", "—", "—")
    else:
        add(4, "UCP Discoverability", "WARN",
            "No SearchAction potentialAction on WebSite schema. This signals to AI agents "
            "that the site supports programmatic search — a prerequisite step toward "
            "Universal Commerce Protocol agent navigation.",
            "Add a WebSite schema block with potentialAction SearchAction and a "
            "urlTemplate pointing to your site search: "
            "\"https://yoursite.com/?s={search_term_string}\"",
            "Low", "Low")

    # Sort: FAIL first, then WARN, then PASS; within each group by priority asc
    order = {"FAIL": 0, "WARN": 1, "INFO": 2, "PASS": 3}
    findings.sort(key=lambda x: (order.get(x["Status"], 9), x["Priority"]))
    return findings


# ---------------------------------------------------------------------------
# EXCEL WRITER  — Tab 1: Raw Data  |  Tab 2: Priority Fix List
# ---------------------------------------------------------------------------

def save_excel(rows, findings, filepath, source_url=""):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required. Run: pip install openpyxl")

    wb = Workbook()

    # ── TAB 1: Raw Data ──────────────────────────────────────────────────────
    ws_raw = wb.active
    ws_raw.title = "Raw Data"

    h_fill  = PatternFill("solid", fgColor="1F4E79")
    h_font  = Font(color="FFFFFF", bold=True, size=11)
    h_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    raw_cols = ["source_url", "syntax", "schema_type", "key", "value"]
    for ci, col in enumerate(raw_cols, 1):
        c = ws_raw.cell(row=1, column=ci, value=col.replace("_", " ").title())
        c.fill = h_fill; c.font = h_font; c.alignment = h_align

    for ri, row in enumerate(rows, 2):
        for ci, col in enumerate(raw_cols, 1):
            c = ws_raw.cell(row=ri, column=ci, value=str(row.get(col, "")))
            c.alignment = Alignment(vertical="top", wrap_text=(ci == 5))

    ws_raw.freeze_panes = "A2"
    ws_raw.auto_filter.ref = ws_raw.dimensions
    for col in ws_raw.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=8)
        ws_raw.column_dimensions[col[0].column_letter].width = min(max_len + 4, 55)
    ws_raw.column_dimensions[get_column_letter(5)].width = 70

    # ── TAB 2: Priority Fix List ─────────────────────────────────────────────
    ws_fix = wb.create_sheet(title="Priority Fix List")

    # Banner
    ws_fix.merge_cells("A1:G1")
    bc = ws_fix["A1"]
    bc.value = "GEO & UCP Readiness — Priority Fix List"
    bc.font = Font(bold=True, size=14, color="FFFFFF")
    bc.fill = PatternFill("solid", fgColor="1F4E79")
    bc.alignment = Alignment(horizontal="center", vertical="center")
    ws_fix.row_dimensions[1].height = 30

    # Summary counts row
    fail_c = sum(1 for f in findings if f["Status"] == "FAIL")
    warn_c = sum(1 for f in findings if f["Status"] == "WARN")
    pass_c = sum(1 for f in findings if f["Status"] == "PASS")
    ws_fix.merge_cells("A2:G2")
    sc = ws_fix["A2"]
    sc.value = (f"Summary:  {fail_c} Critical (FAIL)   {warn_c} Recommended (WARN)"
                f"   {pass_c} Passing   |   Source: {source_url}")
    sc.font = Font(bold=True, size=10, color="FFFFFF")
    sc.fill = PatternFill("solid", fgColor="2E4057")
    sc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_fix.row_dimensions[2].height = 18

    # Date row
    ws_fix.merge_cells("A3:G3")
    dc = ws_fix["A3"]
    dc.value = f"Analyzed: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    dc.font = Font(italic=True, size=9, color="595959")
    dc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_fix.row_dimensions[3].height = 14
    ws_fix.row_dimensions[4].height = 6  # spacer

    # Column headers row 5
    fix_cols   = ["Priority", "Category", "Status", "Issue", "Fix", "Effort", "Impact"]
    fix_widths = [10,          22,          10,       55,      55,    10,       10]

    for ci, (col, w) in enumerate(zip(fix_cols, fix_widths), 1):
        c = ws_fix.cell(row=5, column=ci, value=col)
        c.fill = PatternFill("solid", fgColor="2E75B6")
        c.font = Font(color="FFFFFF", bold=True, size=11)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws_fix.column_dimensions[get_column_letter(ci)].width = w
    ws_fix.row_dimensions[5].height = 22

    status_colors = {
        "FAIL": ("C00000", "FFFFFF"),
        "WARN": ("F4B942", "000000"),
        "PASS": ("375623", "FFFFFF"),
        "INFO": ("2E75B6", "FFFFFF"),
    }
    thin   = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def write_section(start_row, section_findings, label=None):
        r = start_row
        if label and section_findings:
            ws_fix.merge_cells(f"A{r}:G{r}")
            lc = ws_fix.cell(row=r, column=1, value=label)
            lc.fill = PatternFill("solid", fgColor="D6DCE4")
            lc.font = Font(bold=True, size=10, color="1F4E79")
            lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws_fix.row_dimensions[r].height = 16
            r += 1
        for f in section_findings:
            bg, fg = status_colors.get(f["Status"], ("FFFFFF", "000000"))
            for ci, col in enumerate(fix_cols, 1):
                c = ws_fix.cell(row=r, column=ci, value=f.get(col, ""))
                c.border = border
                c.alignment = Alignment(
                    vertical="top", wrap_text=True,
                    horizontal="center" if ci in (1, 3, 6, 7) else "left"
                )
                if ci == 3:   # Status cell — colored badge
                    c.fill = PatternFill("solid", fgColor=bg)
                    c.font = Font(bold=True, color=fg, size=10)
                elif ci == 1: # Priority cell — bold number
                    c.font = Font(bold=True, size=12)
                else:
                    c.font = Font(size=10)
            ws_fix.row_dimensions[r].height = 65
            r += 1
        return r

    action_f = [f for f in findings if f["Status"] in ("FAIL", "WARN", "INFO")]
    pass_f   = [f for f in findings if f["Status"] == "PASS"]

    next_row = write_section(6, action_f)
    next_row += 1
    write_section(next_row, pass_f, "✓  Passing Checks — No Action Required")

    ws_fix.freeze_panes = "A6"

    # Make Priority Fix List the active/first visible sheet
    wb.active = ws_fix

    os.makedirs(os.path.dirname(os.path.abspath(filepath)), exist_ok=True)
    wb.save(filepath)
    return filepath


# ---------------------------------------------------------------------------
# SHARED OUTPUT HELPER
# ---------------------------------------------------------------------------

def _save_output(rows, source_url, filename_stem, output_dir):
    findings = analyze(rows, source_url=source_url)
    fail_c = sum(1 for f in findings if f["Status"] == "FAIL")
    warn_c = sum(1 for f in findings if f["Status"] == "WARN")
    pass_c = sum(1 for f in findings if f["Status"] == "PASS")
    print(f"[scraper] Analysis: {fail_c} FAIL  {warn_c} WARN  {pass_c} PASS")
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename_stem + ".xlsx")
    save_excel(rows, findings, filepath, source_url=source_url)
    print(f"[scraper] Saved → {os.path.abspath(filepath)}")
    return os.path.abspath(filepath)


# ---------------------------------------------------------------------------
# CORE SCRAPER
# ---------------------------------------------------------------------------

@retry(stop=stop_after_attempt(2), wait=wait_exponential(multiplier=1, min=4, max=10))
async def _scrape(url: str) -> dict:
    user_agent = (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/131.0.0.0 Safari/537.36"
    )
    async with async_playwright() as p:
        launch_args = [
            "--disable-blink-features=AutomationControlled",
            "--no-sandbox", "--disable-gpu", "--disable-dev-shm-usage",
        ]
        if os.name != "nt":
            launch_args.extend(["--single-process", "--no-zygote"])

        browser = await p.chromium.launch(headless=True, args=launch_args)
        context = await browser.new_context(
            user_agent=user_agent,
            viewport={"width": 1920, "height": 1080},
            locale="en-US",
        )
        page = await context.new_page()
        await Stealth().apply_stealth_async(page)
        await page.route(
            "**/*.{png,jpg,jpeg,gif,webp,svg,woff,woff2}",
            lambda route: route.abort()
        )
        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=60000)
            await page.wait_for_timeout(5000)
        except Exception as e:
            print(f"[scraper] Navigation warning: {e}", file=sys.stderr)

        page_title = await page.title()
        print(f"[scraper] Page title: {page_title}", file=sys.stderr)

        cf_blocked = any(p in page_title.lower()
                         for p in ("just a moment", "attention required", "access denied"))
        if cf_blocked:
            print("[scraper] Cloudflare detected, waiting 10s…", file=sys.stderr)
            await page.wait_for_timeout(10000)
            page_title = await page.title()
            cf_blocked = any(p in page_title.lower()
                             for p in ("just a moment", "attention required", "access denied"))

        html_content = await page.content()
        current_url  = page.url
        await browser.close()

        if cf_blocked:
            return {"error": "cloudflare_blocked",
                    "message": f"Cloudflare protection detected. Title: '{page_title}'",
                    "title": page_title}

        base_url = get_base_url(html_content, current_url)
        data = extruct.extract(html_content, base_url=base_url, uniform=True,
                               syntaxes=["json-ld", "microdata", "rdfa"])

        if not data.get("json-ld"):
            matches = re.findall(
                r'<script\s+type=["\']application/ld\+json["\']>(.*?)</script>',
                html_content, re.DOTALL | re.IGNORECASE)
            fallback = []
            for m in matches:
                try:
                    parsed = json.loads(m.strip())
                    fallback.extend(parsed if isinstance(parsed, list) else [parsed])
                except Exception:
                    continue
            if fallback:
                data["json-ld"] = fallback

        data["_js_vars_html"] = html_content
        return data


# ---------------------------------------------------------------------------
# PUBLIC ENTRY POINT — URL
# ---------------------------------------------------------------------------

def scrape_to_csv(url: str, output_dir: str = ".") -> str:
    """
    Scrape a live URL, extract structured metadata, analyze GEO/UCP readiness,
    and save a two-tab Excel workbook (.xlsx).

    Tab 1 — Raw Data       : every schema row extracted from the page
    Tab 2 — Priority Fix List : color-coded GEO/UCP findings sorted by severity

    Usage (Jupyter):
        from scraper import scrape_to_csv
        path = scrape_to_csv("https://www.example.com/product/123")
        path = scrape_to_csv("https://www.example.com/product/123", output_dir=r"C:\\Reports")
    """
    # Python 3.14 removed the old get_event_loop() API. This approach works on
    # Python 3.10–3.14+ and inside Jupyter (which has a running event loop).
    import concurrent.futures
    with concurrent.futures.ThreadPoolExecutor(max_workers=1) as pool:
        data = pool.submit(asyncio.run, _scrape(url)).result()

    if "error" in data:
        raise RuntimeError(f"Scraper error: {data.get('message', data['error'])}")

    html_content = data.pop("_js_vars_html", "")
    schema_rows  = flatten_data(data, source_url=url)
    js_rows      = extract_js_variables(html_content, source_url=url)
    rows         = schema_rows + js_rows

    if not rows:
        print("[scraper] WARNING: No data found — bot detection likely served a stripped page.\n"
              "          Try saving the page manually and using scrape_file_to_csv() instead.")
        rows = [{"source_url": url, "syntax": "", "schema_type": "", "key": "", "value": ""}]
    else:
        print(f"[scraper] Found {len(schema_rows)} schema + "
              f"{len(js_rows)} JS rows = {len(rows)} total.")

    return _save_output(rows, url, url_to_filename(url), output_dir)


# ---------------------------------------------------------------------------
# PUBLIC ENTRY POINT — Local File
# ---------------------------------------------------------------------------

def scrape_file_to_csv(file_path: str, url: str = None, output_dir: str = ".") -> str:
    """
    Parse a locally saved HTML file, analyze GEO/UCP readiness,
    and save a two-tab Excel workbook (.xlsx).

    Tab 1 — Raw Data       : every schema row extracted from the file
    Tab 2 — Priority Fix List : color-coded GEO/UCP findings sorted by severity

    Usage (Jupyter):
        from scraper import scrape_file_to_csv

        # Recommended — pass the original URL for clean filename and source_url column
        path = scrape_file_to_csv(
            r"C:\\Users\\You\\Downloads\\page.html",
            url="https://www.example.com/product/123"
        )

        # Without URL — filename derived from the saved file's name
        path = scrape_file_to_csv(r"C:\\Users\\You\\Downloads\\page.html")
    """
    file_path = os.path.abspath(file_path)
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"HTML file not found: {file_path}")

    with open(file_path, "r", encoding="utf-8", errors="replace") as f:
        html_content = f.read()
    print(f"[scraper] Read {len(html_content):,} chars from {os.path.basename(file_path)}")

    source_url = url if url else file_path
    if url:
        filename_stem = url_to_filename(url)
    else:
        base = os.path.splitext(os.path.basename(file_path))[0]
        filename_stem = f"{base}_{datetime.now().strftime('%m%d%Y')}"

    base_url = get_base_url(html_content, source_url)
    data = extruct.extract(html_content, base_url=base_url, uniform=True,
                           syntaxes=["json-ld", "microdata", "rdfa"])

    if not data.get("json-ld"):
        matches = re.findall(
            r'<script\s+type=["\']application/ld\+json["\']>(.*?)</script>',
            html_content, re.DOTALL | re.IGNORECASE)
        fallback = []
        for m in matches:
            try:
                parsed = json.loads(m.strip())
                fallback.extend(parsed if isinstance(parsed, list) else [parsed])
            except Exception:
                continue
        if fallback:
            data["json-ld"] = fallback

    schema_rows = flatten_data(data, source_url=source_url)
    js_rows     = extract_js_variables(html_content, source_url=source_url)
    rows        = schema_rows + js_rows

    if not rows:
        print("[scraper] WARNING: No data found.\n"
              "          Ensure the page was saved as 'Webpage, HTML Only' (not Complete).")
        rows = [{"source_url": source_url, "syntax": "", "schema_type": "", "key": "", "value": ""}]
    else:
        print(f"[scraper] Found {len(schema_rows)} schema + "
              f"{len(js_rows)} JS rows = {len(rows)} total.")

    return _save_output(rows, source_url, filename_stem, output_dir)


# ---------------------------------------------------------------------------
# CLI entry point (original behaviour preserved)
# ---------------------------------------------------------------------------

async def main(url: str):
    data = await _scrape(url)
    data.pop("_js_vars_html", None)
    sys.stdout.write(json.dumps(data))

if __name__ == "__main__":
    if len(sys.argv) > 1:
        asyncio.run(main(sys.argv[1]))
    else:
        sys.stdout.write(json.dumps({"error": "No URL provided"}))
